"""
Jordan Civil Society Organizations Scraper
Source: http://www.civilsociety-jo.net
Output: jordan_organizations.xlsx

Requirements:
    pip install requests beautifulsoup4 openpyxl

Usage:
    python jordan_orgs_scraper.py
"""

import time
import re
import requests
import urllib3
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

BASE = "http://www.civilsociety-jo.net"

CATEGORY_URLS = [
    "/en/organizations/1/employers-unions",
    "/en/organizations/2/employers-associations",
    "/en/organizations/3/chambers-of-commerce",
    "/en/organizations/4/chambers-of-industry",
    "/en/organizations/5/professional-associations",
    "/en/organizations/6/professional-societies",
    "/en/organizations/7/women-organizations",
    "/en/organizations/8/health-care-organizations",
    "/en/organizations/9/youth--sport-clubs",
    "/en/organizations/10/cultural--scientific-organizations",
    "/en/organizations/11/persons-with-disabilities-organizations",
    "/en/organizations/12/environmental-organizations",
    "/en/organizations/13/child--orphas-care-organizations",
    "/en/organizations/14/graduates-of-universities-and-institutes",
    "/en/organizations/15/charities",
    "/en/organizations/16/human-rights-organizations",
    "/en/organizations/17/trade-unions",
    "/en/organizations/18/special-commissions",
    "/en/organizations/19/research-centers",
    "/en/organizations/20/foreign-organizations",
]

CATEGORY_NAMES = {
    "1": "Employers Unions", "2": "Employers Associations",
    "3": "Chambers of Commerce", "4": "Chambers of Industry",
    "5": "Professional Associations", "6": "Professional Societies",
    "7": "Women Organizations", "8": "Health Care Organizations",
    "9": "Youth & Sport Clubs", "10": "Cultural & Scientific Organizations",
    "11": "Persons with Disabilities Organizations", "12": "Environmental Organizations",
    "13": "Child & Orphans Care Organizations", "14": "Graduates of Universities and Institutes",
    "15": "Charities", "16": "Human Rights Organizations",
    "17": "Trade Unions", "18": "Special Commissions",
    "19": "Research Centers", "20": "Foreign Organizations",
}

REQUEST_DELAY = 0.6


def get_session():
    s = requests.Session()
    s.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        )
    })
    return s


def fetch(session, url, retries=3):
    for attempt in range(retries):
        try:
            r = session.get(url, timeout=20, verify=False)
            if r.status_code == 200:
                return r.text
        except requests.RequestException as e:
            if attempt == retries - 1:
                print(f"    Failed: {url} — {e}")
            time.sleep(2)
    return None


def collect_org_urls(session):
    org_urls = {}
    for cat_path in CATEGORY_URLS:
        cat_id = cat_path.split("/")[3]
        cat_name = CATEGORY_NAMES.get(cat_id, "Unknown")
        url = BASE + cat_path
        print(f"  Collecting: {cat_name} ...", end=" ", flush=True)
        html = fetch(session, url)
        if not html:
            print("FAILED")
            continue
        soup = BeautifulSoup(html, "html.parser")
        links = soup.find_all("a", href=re.compile(r"/en/organization/\d+/"))
        found = 0
        for a in links:
            href = a["href"]
            if href not in org_urls:
                org_urls[href] = cat_name
                found += 1
        print(f"{found} orgs found")
        time.sleep(REQUEST_DELAY)
    print(f"\n  Total unique org URLs collected: {len(org_urls)}\n")
    return org_urls


def scrape_org(session, url, category):
    html = fetch(session, url)
    if not html:
        return None
    soup = BeautifulSoup(html, "html.parser")
    h1 = soup.find("h1")
    if not h1:
        return None
    name = h1.get_text(strip=True)
    if not name:
        return None

    full_text = soup.get_text("\n", strip=True)

    website = ""
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if (href.startswith("http") and
                "civilsociety-jo.net" not in href and
                "phenixcenter" not in href and
                "facebook.com" not in href):
            website = href
            break

    email = ""
    mailto = soup.find("a", href=re.compile(r"^mailto:", re.I))
    if mailto:
        email = mailto["href"].replace("mailto:", "").strip()
    else:
        m = re.search(r"[\w.+-]+@[\w-]+\.[a-zA-Z]{2,}", full_text)
        if m:
            email = m.group()

    phone_pat = re.compile(r"(?:\+962[\s\-]?|0)(?:6|7\d)[\s\-]?\d{3,4}[\s\-]?\d{4}")
    phones = list(dict.fromkeys(phone_pat.findall(full_text)))
    phone = " / ".join(phones[:3]) if phones else ""

    fax = ""
    fax_m = re.search(r"[Ff]ax[:\s]*(" + phone_pat.pattern + r")", full_text)
    if fax_m:
        fax = fax_m.group(1)

    governorate = ""
    gov_pat = re.compile(
        r"\b(Amman|Irbid|Zarqa|Mafraq|Ajloun|Jerash|Madaba|Balqa|Karak|Tafileh|Ma.?an|Aqaba)\b", re.I)
    gm = gov_pat.search(full_text[:500])
    if gm:
        governorate = gm.group(1).title()

    address = ""
    addr_m = re.search(r"(?:Jordan|P\.?O\.?\s*Box|Street|St\.|Amman|Irbid)[^\n]{5,80}", full_text)
    if addr_m:
        address = addr_m.group().strip()

    facebook = ""
    fb = soup.find("a", href=re.compile(r"facebook\.com", re.I))
    if fb:
        facebook = fb["href"]

    return {
        "Name": name, "Category": category, "Governorate": governorate,
        "Phone": phone, "Fax": fax, "Email": email, "Website": website,
        "Facebook": facebook, "Address": address, "Source URL": url,
    }


def build_excel(records, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Organizations"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1F4E79")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill("solid", start_color="EBF3FB")

    columns = [
        ("No.", 5), ("Organization Name", 45), ("Category", 30),
        ("Governorate", 14), ("Phone", 24), ("Fax", 18), ("Email", 32),
        ("Website", 35), ("Facebook", 35), ("Address", 40), ("Source URL", 50),
    ]

    ws.row_dimensions[1].height = 30
    for col_idx, (header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    fields = ["_row", "Name", "Category", "Governorate", "Phone", "Fax",
              "Email", "Website", "Facebook", "Address", "Source URL"]

    for row_idx, rec in enumerate(records, start=2):
        rec["_row"] = row_idx - 1
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        ws.row_dimensions[row_idx].height = 18
        for col_idx, field in enumerate(fields, start=1):
            val = rec.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = left_wrap
            cell.border = border
            if fill.fill_type:
                cell.fill = fill

    ws.auto_filter.ref = ws.dimensions

    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Total Organizations"
    ws2["B1"] = len(records)
    ws2["A1"].font = Font(bold=True, name="Arial", size=12)
    ws2["B1"].font = Font(bold=True, name="Arial", size=12)

    ws2["A3"] = "Governorate"
    ws2["B3"] = "Count"
    ws2["A3"].font = Font(bold=True, name="Arial")
    ws2["B3"].font = Font(bold=True, name="Arial")
    gov_counts = {}
    for r in records:
        g = r.get("Governorate") or "Unknown"
        gov_counts[g] = gov_counts.get(g, 0) + 1
    for i, (gov, cnt) in enumerate(sorted(gov_counts.items()), start=4):
        ws2.cell(row=i, column=1, value=gov).font = Font(name="Arial", size=10)
        ws2.cell(row=i, column=2, value=cnt).font = Font(name="Arial", size=10)

    row_offset = len(gov_counts) + 6
    ws2.cell(row=row_offset, column=1, value="Category").font = Font(bold=True, name="Arial")
    ws2.cell(row=row_offset, column=2, value="Count").font = Font(bold=True, name="Arial")
    cat_counts = {}
    for r in records:
        c = r.get("Category") or "Unknown"
        cat_counts[c] = cat_counts.get(c, 0) + 1
    for i, (cat, cnt) in enumerate(sorted(cat_counts.items()), start=row_offset + 1):
        ws2.cell(row=i, column=1, value=cat).font = Font(name="Arial", size=10)
        ws2.cell(row=i, column=2, value=cnt).font = Font(name="Arial", size=10)

    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 10

    wb.save(output_path)
    print(f"\n  Saved {len(records)} organizations -> {output_path}")


def main():
    output_file = "jordan_organizations.xlsx"
    session = get_session()

    print("=" * 60)
    print("  Jordan Civil Society Organizations Scraper")
    print("  Source: civilsociety-jo.net")
    print("=" * 60)

    print("\n[Step 1] Collecting organization URLs from category pages...\n")
    org_urls = collect_org_urls(session)

    if not org_urls:
        print("No URLs collected. Check your internet connection.")
        return

    print(f"[Step 2] Scraping {len(org_urls)} organization pages...\n")
    records = []
    total = len(org_urls)

    for i, (url, category) in enumerate(org_urls.items(), start=1):
        full_url = BASE + url if url.startswith("/") else url
        result = scrape_org(session, full_url, category)
        if result:
            records.append(result)
            gov = result["Governorate"] or "?"
            print(f"  [{i:4d}/{total}] OK  {result['Name'][:50]:<50}  [{gov}]")
        else:
            print(f"  [{i:4d}/{total}] --  (failed: {url[:60]})")

        if i % 100 == 0:
            build_excel(records, output_file)
            print(f"\n  Progress saved ({len(records)} orgs so far)\n")

        time.sleep(REQUEST_DELAY)

    if records:
        build_excel(records, output_file)
    else:
        print("No records scraped.")


if __name__ == "__main__":
    main()